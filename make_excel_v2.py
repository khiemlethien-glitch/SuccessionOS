"""
SuccessionOS_API_Docs.xlsx — v2
Gộp 2 CSV thành 1 file Excel 3 sheet, tối ưu cho backend dev.
  Sheet 1: Hướng dẫn   — legend + cách đọc
  Sheet 2: API List     — danh sách API backend cần làm (bỏ frontend-internal columns)
  Sheet 3: Field Spec   — chi tiết từng field
"""
import csv, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import BORDER_THIN

# ── Màu sắc & style helpers ──────────────────────────────────────────────────
NAVY   = "1E1B4B"
WHITE  = "FFFFFF"
ALT    = "F0F4FF"

SIDE   = Side(style="thin", color="D1D5DB")
THICK  = Side(style="medium", color="9CA3AF")
BDR    = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)

def solid(h): return PatternFill("solid", fgColor=h)

def af(bold=False, size=9, color="111827", name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color)

def align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Trang_thai → (bg, fg, label backend)
STATUS = {
    "KHONG_CAN": ("E5E7EB", "6B7280", "❌ Không cần — Frontend gọi VnR trực tiếp"),
    "DUNG_MOCK":  ("DBEAFE", "1E40AF", "✅ Cần build — Frontend đang dùng mock"),
    "HARDCODE":   ("FEF3C7", "92400E", "⚠️  Xem cột Hành động"),
    "CHUA_BUILD": ("FEE2E2", "991B1B", "🔲 Cần build — Chưa có UI, build khi frontend ready"),
}
PRIORITY = {
    "P0": ("FEE2E2", "991B1B"),
    "P1": ("FEF3C7", "92400E"),
    "P2": ("DBEAFE", "1E40AF"),
    "P3": ("F0FDF4", "166534"),
}
METHOD = {
    "GET":    ("D1FAE5", "065F46"),
    "POST":   ("DBEAFE", "1E40AF"),
    "PUT":    ("FEF3C7", "92400E"),
    "PATCH":  ("EDE9FE", "5B21B6"),
    "DELETE": ("FEE2E2", "991B1B"),
}
PHAN = {
    "REQUEST":  ("EDE9FE", "5B21B6"),
    "RESPONSE": ("D1FAE5", "065F46"),
}
BAT_BUOC = {
    "Y": ("FEE2E2", "991B1B"),
    "O": ("FEF3C7", "92400E"),
    "N": ("F3F4F6", "6B7280"),
}

# Hardcode rows phân loại thủ công dựa trên phân tích code
# True = backend cần build API, False = derived/tính ở frontend
HARDCODE_NEEDS_API = {
    "6":  (True,  "P0", "Cần build — getCurrentUser() đang đọc localStorage, phải có real /me endpoint"),
    "7":  (True,  "P0", "Cần build — Module access control phải đến từ DB, không hardcode trong shell"),
    "10": (True,  "P0", "Cần build — Network graph cần API riêng, không thể derive từ talent list"),
    "11": (True,  "P0", "Cần build — riskFactors cần lưu DB, hiện chỉ T020 có data, các NV khác thiếu"),
    "16": (False, "P1", "Tối ưu sau — Frontend đang derive từ distinct(talent.department). Có API sẽ đúng hơn nhưng chưa urgent"),
    "18": (False, "P1", "Tối ưu sau — Dashboard đang tự tính KPI từ full list. Có /summary giảm payload nhưng không blocking"),
    "19": (True,  "P1", "Cần build — submit() đang tạo ID giả, dữ liệu mất khi F5"),
    "20": (True,  "P1", "Cần build — Admin CRUD chỉ sửa local state, không persist"),
    "21": (True,  "P1", "Cần build — Admin CRUD chỉ xoá local state, không persist"),
    "29": (True,  "P1", "Cần build — Succession plan update chỉ local, không persist"),
    "30": (True,  "P1", "Cần build — Succession plan delete chỉ local, không persist"),
    "44": (True,  "P2", "Cần build — IDP update chỉ local, không persist"),
    "51": (True,  "P2", "Cần build — Chỉ hiển thị mentor name string, chưa có full mentoring info"),
    "54": (False, "P3", "Tối ưu sau — Dashboard gọi 3 API riêng để tính. /kpi là shortcut tối ưu, không blocking"),
    "55": (False, "P3", "Tối ưu sau — topRisk tính client-side từ talents đã fetch. Không blocking"),
    "59": (True,  "P3", "Cần build — 6 users hardcode, mất khi F5"),
    "60": (True,  "P3", "Cần build — User creation chỉ local state"),
    "61": (True,  "P3", "Cần build — User update chỉ local state"),
}

# ── Đọc CSV ──────────────────────────────────────────────────────────────────
def read_csv(path):
    with open(path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    ncols = len(rows[0])
    out = []
    for r in rows:
        if len(r) < ncols: r = r + [''] * (ncols - len(r))
        elif len(r) > ncols: r = r[:ncols-1] + [', '.join(r[ncols-1:])]
        out.append(r)
    return out

map_rows  = read_csv("API_INTEGRATION_MAP.csv")
spec_rows = read_csv("API_FIELD_SPEC.csv")

# Build STT→Trang_thai lookup
map_lookup = {r[0]: r for r in map_rows[1:]}
khong_can_stts = {r[0] for r in map_rows[1:] if r[5] == 'KHONG_CAN'}

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Hướng dẫn
# ════════════════════════════════════════════════════════════════════════════
ws_guide = wb.active
ws_guide.title = "📋 Hướng dẫn"
ws_guide.sheet_view.showGridLines = False
ws_guide.column_dimensions['A'].width = 3
ws_guide.column_dimensions['B'].width = 26
ws_guide.column_dimensions['C'].width = 70
ws_guide.column_dimensions['D'].width = 18

def guide_header(ws, row, text, bg=NAVY, fg=WHITE, size=12):
    ws.row_dimensions[row].height = 28
    c = ws.cell(row=row, column=2, value=text)
    c.fill = solid(bg); c.font = af(bold=True, size=size, color=fg)
    c.alignment = align("left", "center", False)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

def guide_row(ws, row, label, value, label_bg="F3F4F6", label_fg="374151",
              val_bg="FFFFFF", val_fg="111827", height=20):
    ws.row_dimensions[row].height = height
    lc = ws.cell(row=row, column=2, value=label)
    lc.fill = solid(label_bg); lc.font = af(bold=True, size=9, color=label_fg)
    lc.alignment = align("left", "center", False)
    lc.border = BDR
    vc = ws.cell(row=row, column=3, value=value)
    vc.fill = solid(val_bg); vc.font = af(size=9, color=val_fg)
    vc.alignment = align("left", "center", True)
    vc.border = BDR
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

r = 1
# Title
ws_guide.row_dimensions[r].height = 40
tc = ws_guide.cell(row=r, column=2,
    value="SuccessionOS — API Contract cho Backend Developer")
tc.fill = solid(NAVY)
tc.font = af(bold=True, size=14, color=WHITE)
tc.alignment = align("left", "center", False)
ws_guide.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
r += 1

ws_guide.row_dimensions[r].height = 18
sc = ws_guide.cell(row=r, column=2,
    value="Mô tả chính xác từng API và từng field frontend cần — dựa trên phân tích source code thực tế")
sc.font = af(size=9, color="6B7280"); sc.alignment = align("left","center",False)
ws_guide.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
r += 2

# File structure
guide_header(ws_guide, r, "File này gồm 3 sheets", "374151"); r += 1
guide_row(ws_guide, r, "📋  Sheet này",     "Hướng dẫn đọc file + legend màu sắc"); r += 1
guide_row(ws_guide, r, "🔧  API List",       "Danh sách API cần build — đã lọc bỏ những gì frontend tự xử lý"); r += 1
guide_row(ws_guide, r, "📐  Field Spec",     "Chi tiết từng field (request/response) — dùng khi implement endpoint cụ thể"); r += 2

# What to care
guide_header(ws_guide, r, "Backend dev cần đọc gì?", "374151"); r += 1
guide_row(ws_guide, r, "Bước 1 — Nhìn tổng quan",
    "Mở sheet 🔧 API List → lọc cột Hành động ≠ 'Không cần' → đây là việc cần làm"); r += 1
guide_row(ws_guide, r, "Bước 2 — Ưu tiên build",
    "Sắp xếp theo cột Priority: P0 → P1 → P2 → P3"); r += 1
guide_row(ws_guide, r, "Bước 3 — Implement endpoint",
    "Mở sheet 📐 Field Spec → lọc theo Endpoint → xem chính xác từng field cần trả về"); r += 2

# Priority legend
guide_header(ws_guide, r, "Priority — Ưu tiên build", "374151"); r += 1
guide_row(ws_guide, r, "P0 — Ưu tiên cao nhất",
    "Dashboard, Talent list, Profile hoạt động. Không có P0 → app không dùng được",
    *PRIORITY["P0"], *PRIORITY["P0"]); r += 1
guide_row(ws_guide, r, "P1 — Quan trọng",
    "Positions, Succession, Admin CRUD. Thiếu thì dữ liệu mất khi F5",
    *PRIORITY["P1"], *PRIORITY["P1"]); r += 1
guide_row(ws_guide, r, "P2 — Trung bình",
    "IDP, Mentoring, Assessment. Module đang disabled trong UI",
    *PRIORITY["P2"], *PRIORITY["P2"]); r += 1
guide_row(ws_guide, r, "P3 — Thấp",
    "Dashboard KPI shortcut, Users admin, Reports. Hoạt động được không cần P3",
    *PRIORITY["P3"], *PRIORITY["P3"]); r += 2

# Trang_thai legend
guide_header(ws_guide, r, "Cột 'Hành động' — ý nghĩa", "374151"); r += 1
guide_row(ws_guide, r, "✅ Cần build — Frontend đang mock",
    "Frontend đang đọc file JSON tĩnh. Khi backend xong → frontend flip useMock: false là chạy",
    "DBEAFE", "1E40AF", "DBEAFE", "1E40AF"); r += 1
guide_row(ws_guide, r, "⚠️  Cần build — Frontend đang hardcode",
    "Data đang gán cứng trong code .ts. Phải có API thật để dữ liệu persist và đúng theo user",
    "FEF3C7", "92400E", "FEF3C7", "92400E"); r += 1
guide_row(ws_guide, r, "📊 Tối ưu sau",
    "Frontend tự tính được từ data đã fetch. Có API sẽ tốt hơn nhưng không blocking — làm sau P1",
    "F0FDF4", "166534", "F0FDF4", "166534"); r += 1
guide_row(ws_guide, r, "🔲 Chờ frontend",
    "Chưa có màn hình UI. Build khi frontend báo ready",
    "F3F4F6", "374151", "F3F4F6", "374151"); r += 1
guide_row(ws_guide, r, "❌ Không cần",
    "Frontend gọi trực tiếp VnR HRM OIDC. Backend không cần implement",
    "E5E7EB", "6B7280", "E5E7EB", "6B7280"); r += 2

# Field Spec guide
guide_header(ws_guide, r, "Cách đọc sheet Field Spec", "374151"); r += 1
guide_row(ws_guide, r, "Phan = REQUEST",  "Field nằm trong request body / query param frontend gửi lên",   "EDE9FE","5B21B6","EDE9FE","5B21B6"); r += 1
guide_row(ws_guide, r, "Phan = RESPONSE", "Field frontend expect trong response — thiếu field → UI bị lỗi","D1FAE5","065F46","D1FAE5","065F46"); r += 1
guide_row(ws_guide, r, "Bat_buoc = Y",    "Bắt buộc. Thiếu → frontend crash hoặc hiển thị sai",           "FEE2E2","991B1B","FEE2E2","991B1B"); r += 1
guide_row(ws_guide, r, "Bat_buoc = O",    "Optional. Thiếu → UI fallback về giá trị mặc định",            "FEF3C7","92400E","FEF3C7","92400E"); r += 1
guide_row(ws_guide, r, "Logic_xu_ly",     "Mô tả chính xác cách frontend dùng field — dùng để validate response đúng không","F9FAFB","374151"); r += 1

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — API List (dành cho backend dev)
# ════════════════════════════════════════════════════════════════════════════
ws_api = wb.create_sheet("🔧 API List")
ws_api.sheet_view.showGridLines = False
ws_api.freeze_panes = "A2"

# Columns backend dev cần (bỏ frontend-internal cols)
# MAP: STT Module Method Endpoint Mo_ta Trang_thai File_component Ham_can_sua Mock_hien_tai Request_params Response_fields_can Priority Ghi_chu
# idx:  0    1      2      3        4      5           6              7          8              9              10                  11       12

OUT_HEADERS = [
    "Priority", "Hành động",
    "Module", "Method", "Endpoint", "Mô tả",
    "Request params", "Response fields", "Ghi chú cho backend"
]

def get_hanh_dong(row):
    stt, status, priority, note = row[0], row[5], row[11], row[12]
    if status == "KHONG_CAN":
        return ("❌ Không cần — Frontend gọi VnR trực tiếp", "E5E7EB", "6B7280")
    if status == "DUNG_MOCK":
        return ("✅ Cần build — Frontend đang dùng mock", "DBEAFE", "1E40AF")
    if status == "CHUA_BUILD":
        return ("🔲 Chờ frontend — Build khi UI ready", "F3F4F6", "374151")
    if status == "HARDCODE":
        info = HARDCODE_NEEDS_API.get(stt)
        if info:
            needs_api, _, detail = info
            if needs_api:
                return ("⚠️  Cần build — Frontend đang hardcode", "FEF3C7", "92400E")
            else:
                return ("📊 Tối ưu sau — Frontend tự tính được", "F0FDF4", "166534")
        return ("⚠️  Cần build — Frontend đang hardcode", "FEF3C7", "92400E")
    return ("", "FFFFFF", "111827")

# Sort key: P0 first, then by trang_thai priority
STATUS_ORDER = {"KHONG_CAN": 4, "DUNG_MOCK": 0, "HARDCODE": 1, "CHUA_BUILD": 2}
PRIO_ORDER   = {"P0": 0, "P1": 1, "P2": 2, "P3": 3, "": 9}

data_rows = map_rows[1:]
data_rows_sorted = sorted(
    data_rows,
    key=lambda r: (PRIO_ORDER.get(r[11], 9), STATUS_ORDER.get(r[5], 9))
)

# Write header
ws_api.row_dimensions[1].height = 36
for ci, h in enumerate(OUT_HEADERS, 1):
    c = ws_api.cell(row=1, column=ci, value=h)
    c.fill = solid(NAVY); c.font = af(bold=True, size=10, color=WHITE)
    c.alignment = align("center", "center", True); c.border = BDR

ws_api.auto_filter.ref = f"A1:{get_column_letter(len(OUT_HEADERS))}1"

# Write data
for ri, row in enumerate(data_rows_sorted, start=2):
    stt, status, priority = row[0], row[5], row[11]
    hanh_dong, hd_bg, hd_fg = get_hanh_dong(row)

    # Get better Ghi_chu: merge original + HARDCODE detail if exists
    ghi_chu = row[12]
    if status == "HARDCODE" and stt in HARDCODE_NEEDS_API:
        _, _, detail = HARDCODE_NEEDS_API[stt]
        ghi_chu = detail

    out_vals = [
        priority,           # Priority
        hanh_dong,          # Hành động
        row[1],             # Module
        row[2],             # Method
        row[3],             # Endpoint
        row[4],             # Mô tả
        row[9],             # Request params
        row[10],            # Response fields
        ghi_chu,            # Ghi chú
    ]

    is_alt = ri % 2 == 0
    base_bg = ALT if is_alt else "FFFFFF"

    for ci, val in enumerate(out_vals, 1):
        c = ws_api.cell(row=ri, column=ci, value=val)
        hdr = OUT_HEADERS[ci-1]
        c.border = BDR
        c.alignment = align("left", "top", True)

        if hdr == "Priority":
            bg, fg = PRIORITY.get(val, ("FFFFFF", "111827"))
            c.fill = solid(bg); c.font = af(bold=True, size=9, color=fg)
            c.alignment = align("center", "top", False)
        elif hdr == "Hành động":
            c.fill = solid(hd_bg); c.font = af(bold=False, size=9, color=hd_fg)
        elif hdr == "Method":
            bg, fg = METHOD.get(val.upper(), ("FFFFFF", "111827"))
            c.fill = solid(bg); c.font = af(bold=True, size=9, color=fg)
            c.alignment = align("center", "top", False)
        else:
            c.fill = solid(base_bg); c.font = af(size=9)

# Column widths for API List
COL_W = [9, 38, 18, 9, 42, 30, 32, 40, 60]
for i, w in enumerate(COL_W, 1):
    ws_api.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Field Spec
# ════════════════════════════════════════════════════════════════════════════
ws_spec = wb.create_sheet("📐 Field Spec")
ws_spec.sheet_view.showGridLines = False
ws_spec.freeze_panes = "A2"

# Add Priority column (lookup from MAP)
SPEC_HEADERS = ["Priority", "Hành động"] + spec_rows[0]

ws_spec.row_dimensions[1].height = 36
for ci, h in enumerate(SPEC_HEADERS, 1):
    c = ws_spec.cell(row=1, column=ci, value=h)
    c.fill = solid(NAVY); c.font = af(bold=True, size=10, color=WHITE)
    c.alignment = align("center", "center", True); c.border = BDR

ws_spec.auto_filter.ref = f"A1:{get_column_letter(len(SPEC_HEADERS))}1"

# Filter out KHONG_CAN rows, add priority + hanh_dong from MAP lookup
spec_data = [r for r in spec_rows[1:] if r[0] not in khong_can_stts]

# Sort by priority
spec_data_sorted = sorted(
    spec_data,
    key=lambda r: (PRIO_ORDER.get(map_lookup.get(r[0], ['','','','','','','','','','','','','',''])[11], 9),
                   r[0], r[3])  # priority, STT_API, Phan (REQUEST before RESPONSE)
)

for ri, row in enumerate(spec_data_sorted, start=2):
    stt_api = row[0]
    map_row = map_lookup.get(stt_api)
    priority  = map_row[11] if map_row else ""
    hanh_dong, hd_bg, hd_fg = get_hanh_dong(map_row) if map_row else ("","FFFFFF","111827")

    full_row = [priority, hanh_dong] + row
    is_alt = ri % 2 == 0
    base_bg = ALT if is_alt else "FFFFFF"

    for ci, val in enumerate(full_row, 1):
        c = ws_spec.cell(row=ri, column=ci, value=val)
        hdr = SPEC_HEADERS[ci-1]
        c.border = BDR
        c.alignment = align("left", "top", True)

        if hdr == "Priority":
            bg, fg = PRIORITY.get(val, ("FFFFFF","111827"))
            c.fill = solid(bg); c.font = af(bold=True, size=9, color=fg)
            c.alignment = align("center","top",False)
        elif hdr == "Hành động":
            c.fill = solid(hd_bg); c.font = af(size=9, color=hd_fg)
        elif hdr == "Method":
            bg, fg = METHOD.get(val.upper(), ("FFFFFF","111827"))
            c.fill = solid(bg); c.font = af(bold=True, size=9, color=fg)
            c.alignment = align("center","top",False)
        elif hdr == "Phan":
            bg, fg = PHAN.get(val.upper(), ("FFFFFF","111827"))
            c.fill = solid(bg); c.font = af(bold=True, size=9, color=fg)
            c.alignment = align("center","top",False)
        elif hdr == "Bat_buoc":
            bg, fg = BAT_BUOC.get(val.upper(), ("FFFFFF","111827"))
            c.fill = solid(bg); c.font = af(bold=True, size=9, color=fg)
            c.alignment = align("center","top",False)
        else:
            c.fill = solid(base_bg); c.font = af(size=9)

# Column widths for Field Spec
# SPEC_HEADERS: Priority Hành_động STT_API Method Endpoint Phan Ten_field Kieu Bat_buoc Gia_tri_vi_du Range Mo_ta_cd Vi_tri_UI Component Ham_su_dung Logic
SPEC_COL_W = [9, 36, 8, 9, 38, 11, 32, 10, 10, 22, 20, 44, 38, 44, 28, 50]
for i, w in enumerate(SPEC_COL_W, 1):
    ws_spec.column_dimensions[get_column_letter(i)].width = w

# ── Save ──────────────────────────────────────────────────────────────────────
out = "SuccessionOS_API_Docs.xlsx"
wb.save(out)

# Stats
n_need  = sum(1 for r in data_rows if r[5] != "KHONG_CAN")
n_p0    = sum(1 for r in data_rows if r[11] == "P0" and r[5] != "KHONG_CAN")
n_mock  = sum(1 for r in data_rows if r[5] == "DUNG_MOCK")
n_hc    = sum(1 for r in data_rows if r[5] == "HARDCODE")
n_spec  = len(spec_data_sorted)

print(f"✅  Saved: {out}")
print(f"   Sheet 1 — Hướng dẫn : legend + cách đọc")
print(f"   Sheet 2 — API List  : {n_need} API cần làm (bỏ 5 KHONG_CAN)")
print(f"             P0={n_p0} | DUNG_MOCK={n_mock} | HARDCODE={n_hc}")
print(f"   Sheet 3 — Field Spec: {n_spec} field rows (KHONG_CAN đã lọc)")
