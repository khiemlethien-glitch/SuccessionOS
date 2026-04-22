"""
Tạo file SuccessionOS_API_Docs.xlsx từ 2 CSV:
  - Sheet 1: API_INTEGRATION_MAP (61 API)
  - Sheet 2: API_FIELD_SPEC (per-field detail)
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

# ─── Màu sắc ─────────────────────────────────────────────────────────────────
HEADER_BG   = "1E1B4B"   # Navy (brand)
HEADER_FG   = "FFFFFF"
ROW_ALT     = "F0F4FF"   # Xanh nhạt xen kẽ

STATUS_COLORS = {
    "KHONG_CAN": ("FEF3C7", "92400E"),   # Vàng nhạt / nâu
    "DUNG_MOCK":  ("DBEAFE", "1E40AF"),  # Xanh dương nhạt / đậm
    "HARDCODE":   ("FEE2E2", "991B1B"),  # Đỏ nhạt / đậm
    "CHUA_BUILD": ("F3F4F6", "374151"),  # Xám nhạt / đậm
}

PRIORITY_COLORS = {
    "P0": ("FEE2E2", "991B1B"),
    "P1": ("FEF3C7", "92400E"),
    "P2": ("DBEAFE", "1E40AF"),
    "P3": ("F0FDF4", "166534"),
}

SIDE = Side(style="thin", color="D1D5DB")
BORDER = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)


def header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def apply_header(ws, col_count: int):
    """Format hàng 1 (header)."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill(HEADER_BG)
        cell.font = Font(bold=True, color=HEADER_FG, size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def auto_width(ws, min_w=8, max_w=50):
    """Tự chỉnh độ rộng cột theo nội dung."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))


# ─── Đọc CSV ─────────────────────────────────────────────────────────────────
map_df  = pd.read_csv("/Users/lethienkhiem/Documents/CodeProject/successionos/API_INTEGRATION_MAP.csv",
                      dtype=str, keep_default_na=False)
spec_df = pd.read_csv("/Users/lethienkhiem/Documents/CodeProject/successionos/API_FIELD_SPEC.csv",
                      dtype=str, keep_default_na=False)

# ─── Xuất sang Excel ─────────────────────────────────────────────────────────
out_path = "/Users/lethienkhiem/Documents/CodeProject/successionos/SuccessionOS_API_Docs.xlsx"

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    map_df.to_excel(writer,  sheet_name="API_Integration_Map", index=False)
    spec_df.to_excel(writer, sheet_name="API_Field_Spec",      index=False)

wb = load_workbook(out_path)

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — API_Integration_Map
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb["API_Integration_Map"]

apply_header(ws1, len(map_df.columns))
ws1.row_dimensions[1].height = 36

# Freeze header
ws1.freeze_panes = "A2"

# Thêm AutoFilter
ws1.auto_filter.ref = ws1.dimensions

# Tìm cột Trang_thai và Priority
cols = {cell.value: cell.column for cell in ws1[1]}
col_status   = cols.get("Trang_thai")
col_priority = cols.get("Priority")
col_method   = cols.get("Method")

for row_idx, row in enumerate(ws1.iter_rows(min_row=2), start=2):
    is_alt = (row_idx % 2 == 0)

    for cell in row:
        cell.font      = Font(size=9, name="Calibri")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border    = BORDER

        # Màu nền xen kẽ (default)
        if is_alt:
            cell.fill = PatternFill("solid", fgColor=ROW_ALT)

    # Tô màu ô Trang_thai
    if col_status:
        status_cell = ws1.cell(row=row_idx, column=col_status)
        status_val  = (status_cell.value or "").strip()
        if status_val in STATUS_COLORS:
            bg, fg = STATUS_COLORS[status_val]
            status_cell.fill = PatternFill("solid", fgColor=bg)
            status_cell.font = Font(bold=True, color=fg, size=9, name="Calibri")
            status_cell.alignment = Alignment(horizontal="center", vertical="top")

    # Tô màu ô Priority
    if col_priority:
        pri_cell = ws1.cell(row=row_idx, column=col_priority)
        pri_val  = (pri_cell.value or "").strip()
        if pri_val in PRIORITY_COLORS:
            bg, fg = PRIORITY_COLORS[pri_val]
            pri_cell.fill = PatternFill("solid", fgColor=bg)
            pri_cell.font = Font(bold=True, color=fg, size=9, name="Calibri")
            pri_cell.alignment = Alignment(horizontal="center", vertical="top")

    # In đậm Method (GET/POST/PUT/DELETE)
    if col_method:
        m_cell = ws1.cell(row=row_idx, column=col_method)
        m_val  = (m_cell.value or "").strip().upper()
        method_colors = {
            "GET":    ("D1FAE5", "065F46"),
            "POST":   ("DBEAFE", "1E40AF"),
            "PUT":    ("FEF3C7", "92400E"),
            "PATCH":  ("EDE9FE", "5B21B6"),
            "DELETE": ("FEE2E2", "991B1B"),
        }
        if m_val in method_colors:
            bg, fg = method_colors[m_val]
            m_cell.fill = PatternFill("solid", fgColor=bg)
            m_cell.font = Font(bold=True, color=fg, size=9, name="Calibri")
            m_cell.alignment = Alignment(horizontal="center", vertical="top")

# Độ rộng cột cố định theo ngữ nghĩa
col_widths_1 = {
    "STT": 5, "Module": 16, "Method": 9, "Endpoint": 38, "Mo_ta": 32,
    "Trang_thai": 14, "File_component": 50, "Ham_can_sua": 28,
    "Mock_hien_tai": 24, "Request_params": 30, "Response_fields_can": 38,
    "Priority": 9, "Ghi_chu": 55,
}
for hdr_cell in ws1[1]:
    key = hdr_cell.value
    if key in col_widths_1:
        ws1.column_dimensions[get_column_letter(hdr_cell.column)].width = col_widths_1[key]

ws1.sheet_view.showGridLines = False

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — API_Field_Spec
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb["API_Field_Spec"]

apply_header(ws2, len(spec_df.columns))
ws2.row_dimensions[1].height = 36
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

# Tìm cột Phan (REQUEST / RESPONSE)
cols2 = {cell.value: cell.column for cell in ws2[1]}
col_phan     = cols2.get("Phan")
col_bat_buoc = cols2.get("Bat_buoc")
col_method2  = cols2.get("Method")

PHAN_COLORS = {
    "REQUEST":  ("EDE9FE", "5B21B6"),
    "RESPONSE": ("D1FAE5", "065F46"),
}

for row_idx, row in enumerate(ws2.iter_rows(min_row=2), start=2):
    is_alt = (row_idx % 2 == 0)

    for cell in row:
        cell.font      = Font(size=9, name="Calibri")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border    = BORDER
        if is_alt:
            cell.fill = PatternFill("solid", fgColor=ROW_ALT)

    # Tô màu Phan
    if col_phan:
        p_cell = ws2.cell(row=row_idx, column=col_phan)
        p_val  = (p_cell.value or "").strip().upper()
        if p_val in PHAN_COLORS:
            bg, fg = PHAN_COLORS[p_val]
            p_cell.fill = PatternFill("solid", fgColor=bg)
            p_cell.font = Font(bold=True, color=fg, size=9, name="Calibri")
            p_cell.alignment = Alignment(horizontal="center", vertical="top")

    # Tô màu Bat_buoc
    if col_bat_buoc:
        b_cell = ws2.cell(row=row_idx, column=col_bat_buoc)
        b_val  = (b_cell.value or "").strip().upper()
        if b_val == "Y":
            b_cell.fill = PatternFill("solid", fgColor="FEE2E2")
            b_cell.font = Font(bold=True, color="991B1B", size=9, name="Calibri")
            b_cell.alignment = Alignment(horizontal="center", vertical="top")
        elif b_val == "O":
            b_cell.fill = PatternFill("solid", fgColor="FEF3C7")
            b_cell.font = Font(bold=True, color="92400E", size=9, name="Calibri")
            b_cell.alignment = Alignment(horizontal="center", vertical="top")
        elif b_val == "N":
            b_cell.fill = PatternFill("solid", fgColor="F3F4F6")
            b_cell.font = Font(color="6B7280", size=9, name="Calibri")
            b_cell.alignment = Alignment(horizontal="center", vertical="top")

    # Tô màu Method
    if col_method2:
        m_cell = ws2.cell(row=row_idx, column=col_method2)
        m_val  = (m_cell.value or "").strip().upper()
        method_colors2 = {
            "GET":    ("D1FAE5", "065F46"),
            "POST":   ("DBEAFE", "1E40AF"),
            "PUT":    ("FEF3C7", "92400E"),
            "PATCH":  ("EDE9FE", "5B21B6"),
            "DELETE": ("FEE2E2", "991B1B"),
        }
        if m_val in method_colors2:
            bg, fg = method_colors2[m_val]
            m_cell.fill = PatternFill("solid", fgColor=bg)
            m_cell.font = Font(bold=True, color=fg, size=9, name="Calibri")
            m_cell.alignment = Alignment(horizontal="center", vertical="top")

# Độ rộng cột cố định
col_widths_2 = {
    "STT_API": 8, "Method": 9, "Endpoint": 36, "Phan": 11, "Ten_field": 34,
    "Kieu": 10, "Bat_buoc": 10, "Gia_tri_vi_du": 24, "Range_Cho_phep": 20,
    "Mo_ta_cong_dung": 45, "Vi_tri_UI": 40, "Component_file": 48,
    "Ham_su_dung": 30, "Logic_xu_ly": 50,
}
for hdr_cell in ws2[1]:
    key = hdr_cell.value
    if key in col_widths_2:
        ws2.column_dimensions[get_column_letter(hdr_cell.column)].width = col_widths_2[key]

ws2.sheet_view.showGridLines = False

# ─── Lưu ─────────────────────────────────────────────────────────────────────
wb.save(out_path)
print(f"✅ Đã tạo: {out_path}")
print(f"   Sheet 1 — API_Integration_Map: {len(map_df)} dòng")
print(f"   Sheet 2 — API_Field_Spec:      {len(spec_df)} dòng")
